import time
import concurrent.futures
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from constants import driver_path, lok_sabha_num, state_codes, url
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def save_tables_to_excel(data, filename):
    wb = Workbook()
    
    params = ['Seats', 'Votes %', 'Contested Voteshare']
    
    for param in params:
        ws = wb.create_sheet(title=param)
        ws.append(["Region", "NDA", "INDIA"])
        
        for region, values in data.items():
            headers = values[0]
            try:
                idx = headers.index(param)
            except ValueError:
                continue 
            
            nda_value = values[1][idx]
            india_value = values[2][idx]
            
            ws.append([region, nda_value, india_value])
        
        for col_num, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_num)
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
    
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    
    wb.save(filename)


def print_tables(data):
    params = ['Seats', 'Votes %', 'Contested Voteshare']
    
    for param in params:
        table = PrettyTable()
        table.field_names = ["Region", "NDA", "INDIA"]
        table.align["Region"] = "l"
        table.align["NDA"] = "r"
        table.align["INDIA"] = "r"
        
        print(f"\n{param}\n" + "-"*40)
        
        for region, values in data.items():
            headers = values[0]
            try:
                idx = headers.index(param)
            except ValueError:
                continue
            
            nda_value = values[1][idx]
            india_value = values[2][idx]
            
            print(f"{region:<25} | {nda_value:>5} | {india_value:>5}")
        
        print("\n") 


def evaluate_years(year, states):
    for i in range(len(states)):
        if year <= 1999:
            if states[i] == 'Bihar':
                states[i] = 'Bihar [1947 - 1999]'
            elif states[i] == 'Madhya Pradesh':
                states[i] = 'Madhya Pradesh [1947 - 1999]'
            elif states[i] == 'Uttar Pradesh':
                states[i] = 'Uttar Pradesh [1947 - 1999]'
        else:
            if states[i] == 'Bihar':
                states[i] = 'Bihar [2000 Onwards]'
            elif states[i] == 'Madhya Pradesh':
                states[i] = 'Madhya Pradesh [2000 Onwards]'
            elif states[i] == 'Uttar Pradesh':
                states[i] = 'Uttar Pradesh [2000 Onwards]'
        if year >= 1977 and states[i] == 'Delhi':
            states[i] = 'Delhi [1977 Onwards]'

def parse_state(state, year, chrome_options, service):
    driver = None
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options)
        addr = f'{url}/{lok_sabha_num[year]}/{state_codes[state]}'
        driver.get(addr)

        wait = WebDriverWait(driver, 20)
        chart_div = wait.until(
            EC.presence_of_element_located((By.ID, 'charttable_alliance'))
        )
        table = chart_div.find_element(By.CLASS_NAME, 'grid')
        rows = table.find_elements(By.TAG_NAME, 'tr')
        state_data = []
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, 'td') or row.find_elements(By.TAG_NAME, 'th')
            row_texts = [cell.text for cell in cells]
            state_data.append(row_texts)
        return state, state_data
    except Exception as e:
        print(f"Error processing {state}: {e}")
        return state, []
    finally:
        if driver:
            driver.quit()

def main():
    year = 0
    while year < 1947 or year > 2024:
        year = int(input('Enter year: '))
    state_list = input('Enter states (sep by comma): ').split(', ')
    
    chrome_options = Options()
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("--headless")
    service = Service(executable_path=driver_path)

    data = {}

    evaluate_years(year, state_list)
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(parse_state, state, year, chrome_options, service) for state in state_list]
        
        for future in concurrent.futures.as_completed(futures):
            state, state_data = future.result()
            data[state] = state_data

    print_tables(data)
    save_tables_to_excel(data, "election_data.xlsx")
    end_time = time.time()
    print(f"Время выполнения: {end_time - start_time:.2f} секунд")


if __name__ == '__main__':
    main()
